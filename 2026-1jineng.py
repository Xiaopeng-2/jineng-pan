import os
import time
from datetime import datetime
from typing import List, Tuple

# 先设置pandas配置，避免版本兼容问题
import pandas as pd

pd.set_option('io.excel.xlsx.reader', 'openpyxl')  # 强制指定xlsx读取引擎
pd.set_option('io.excel.xls.reader', 'xlrd')  # 兼容xls格式
import streamlit as st
from streamlit_autorefresh import st_autorefresh
from streamlit_echarts import st_echarts
import plotly.graph_objects as go
from openpyxl import load_workbook
import psutil

# -------------------- 页面配置 --------------------
st.set_page_config(page_title="技能覆盖分析大屏", layout="wide")

# -------------------- 页面样式 --------------------
PAGE_CSS = """
<style>
/* 全局主体背景 + 文字颜色 */
body, [data-testid="stAppViewContainer"]{
    background-color: #e6f7ff !important;
    color: #003366 !important;
}

/* 侧边栏背景 + 文字 */
[data-testid="stSidebar"]{
    background-color: #d1e7f5 !important;
    color: #003366 !important;
}

/* 按钮样式 */
div.stButton>button{
    background-color: #4cc9f0 !important;
    color: #000000 !important;
    border-radius:10px;
    height:40px;
    font-weight:700;
    margin:5px 0;
    width:100%;
}
div.stButton>button:hover{
    background-color:#4895ef !important;
    color:#ffffff !important;
}

/* 自定义指标卡片 */
.metric-card{
    background-color: #ffffff !important;
    padding:20px;
    border-radius:16px;
    text-align:center;
    box-shadow:0 0 15px rgba(0,0,0,0.08);
}
.metric-value{
    font-size:36px;
    font-weight:800;
    color: #0066cc !important;
}
.metric-label{
    font-size:14px;
    color: #336699 !important;
}

/* 分割线 */
hr{
    border:none;
    border-top:1px solid #bbd9f7;
    margin:16px 0;
}

/* 热力图滚动容器 */
.heatmap-container {
    max-height: 700px;
    overflow-y: auto;
    overflow-x: auto;
    border-radius: 8px;
    background-color: #ffffff;
}

/* 滚动条美化 */
.heatmap-container::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}
.heatmap-container::-webkit-scrollbar-thumb {
    background-color: #99c2ff;
    border-radius: 4px;
}
.heatmap-container::-webkit-scrollbar-track {
    background-color: #e6f7ff;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

SAVE_FILE = "jixiao.xlsx"   # 固定保存的文件
# 安全校验：确保文件后缀是xlsx，避免格式识别错误
if not SAVE_FILE.endswith(".xlsx"):
    SAVE_FILE += ".xlsx"


# -------------------- 工具函数：检测文件是否被占用 --------------------
def is_file_locked(file_path):
    if not os.path.exists(file_path):
        return False
    try:
        with open(file_path, 'rb+'):
            return False
    except PermissionError:
        return True
    except Exception:
        return True


# -------------------- 数据导入（核心修复：指定读取引擎） --------------------
@st.cache_data(ttl=300)
def load_sheets(file) -> Tuple[List[str], dict]:
    """读取Excel所有工作表，修复pandas版本兼容问题"""
    if not os.path.exists(file):
        return [], {}

    try:
        # 核心修复1：显式指定引擎为openpyxl，避免自动检测出错
        xpd = pd.ExcelFile(file, engine="openpyxl")
    except Exception as e:
        st.sidebar.error(f"❌ 读取Excel文件失败：{e}")
        # 降级尝试xlrd引擎（兼容旧格式）
        try:
            xpd = pd.ExcelFile(file, engine="xlrd")
        except Exception as e2:
            st.sidebar.error(f"❌ xlrd引擎也读取失败：{e2}")
            return [], {}

    frames = {}
    for s in xpd.sheet_names:
        try:
            # 核心修复2：读取时指定引擎，避免格式识别错误
            df0 = pd.read_excel(xpd, sheet_name=s, engine="openpyxl")
            if df0.empty:
                continue
            if not {"明细", "员工", "值"}.issubset(df0.columns):
                st.sidebar.warning(f"⚠️ 表 {s} 缺少必要列，已跳过。")
                continue

            # 解析分组行
            if df0.iloc[0, 0] == "分组":
                groups = df0.iloc[0, 1:].tolist()
                df0 = df0.drop(0).reset_index(drop=True)
                emp_cols = [c for c in df0.columns if c not in ["明细", "数量总和", "编号"]]
                group_map = {emp: groups[i] if i < len(groups) else None for i, emp in enumerate(emp_cols)}
                df_long = df0.melt(
                    id_vars=["明细", "数量总和"] if "数量总和" in df0.columns else ["明细"],
                    value_vars=emp_cols,
                    var_name="员工",
                    value_name="值"
                )
                df_long["分组"] = df_long["员工"].map(group_map)
                frames[s] = df_long
            else:
                frames[s] = df0
        except Exception as e:
            st.sidebar.error(f"❌ 读取 {s} 时出错: {e}")
    return xpd.sheet_names, frames


# -------------------- 优化后的删除工作表函数 --------------------
def delete_sheet_optimized(file_path, sheet_name):
    if not os.path.exists(file_path):
        return False, "❌ 文件不存在"

    if is_file_locked(file_path):
        return False, "❌ 文件被占用（可能Excel已打开），请关闭Excel后重试"

    try:
        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            wb.close()
            return False, "❌ 工作表不存在"

        wb.remove(wb[sheet_name])
        wb.save(file_path)
        wb.close()

        return True, f"✅ 成功删除工作表: {sheet_name}"
    except PermissionError:
        return False, "❌ 权限不足，无法删除工作表（请检查文件是否只读）"
    except Exception as e:
        return False, f"❌ 删除失败: {str(e)}"


# -------------------- 文件读取 --------------------
sheets, sheet_frames = load_sheets(SAVE_FILE)

# 初始化：文件不存在时创建空文件
if not os.path.exists(SAVE_FILE):
    try:
        # 核心修复3：创建文件时显式指定引擎
        with pd.ExcelWriter(SAVE_FILE, engine="openpyxl") as writer:
            pd.DataFrame(columns=["明细", "数量总和", "员工", "值", "分组"]).to_excel(
                writer, sheet_name="示例_2025_01", index=False
            )
        sheets, sheet_frames = load_sheets(SAVE_FILE)
        st.sidebar.success(f"✅ 已创建初始文件 {SAVE_FILE}")
    except Exception as e:
        st.sidebar.error(f"❌ 创建初始文件失败：{e}")
elif not sheets:
    st.sidebar.warning("⚠️ 文件存在但无有效工作表，已创建示例数据")
    sheet_frames = {
        "示例_2025_01": pd.DataFrame({
            "明细": ["任务A", "任务B", "任务C"],
            "数量总和": [3, 2, 5],
            "员工": ["张三", "李四", "王五"],
            "值": [1, 1, 1],
            "分组": ["A8", "B7", "VN"]
        })
    }
    sheets = ["示例_2025_01"]
else:
    st.sidebar.success(f"✅ 已加载库文件 {SAVE_FILE}（共{len(sheets)}个工作表）")

# ---------- 🧠 自动检测并修复数量总和 ----------
repaired_count = 0
repaired_frames = {}
for sheet_name, df0 in sheet_frames.items():
    if "明细" in df0.columns and "值" in df0.columns:
        if "数量总和" not in df0.columns or df0["数量总和"].isnull().any():
            repaired = True
        else:
            true_sum = df0.groupby("明细")["值"].sum().reset_index()
            merged = df0.merge(true_sum, on="明细", how="left", suffixes=("", "_真实"))
            repaired = not merged["数量总和"].equals(merged["值_真实"])

        if repaired:
            repaired_count += 1
            sum_df = (
                df0.groupby("明细", as_index=False)["值"].sum()
                .rename(columns={"值": "数量总和"})
            )
            df0 = df0.drop(columns=["数量总和"], errors="ignore")
            df0 = df0.merge(sum_df, on="明细", how="left")
            repaired_frames[sheet_name] = df0

if repaired_frames:
    try:
        with pd.ExcelWriter(SAVE_FILE, engine="openpyxl") as writer:
            for sn in sheets:
                if sn in repaired_frames:
                    repaired_df = repaired_frames[sn]
                    repaired_df.to_excel(writer, sheet_name=sn, index=False)
                    sheet_frames[sn] = repaired_df
                else:
                    df_original = pd.read_excel(SAVE_FILE, sheet_name=sn, engine="openpyxl")
                    df_original.to_excel(writer, sheet_name=sn, index=False)
        st.cache_data.clear()
        st.sidebar.info(f"🔧 已自动修复 {repaired_count} 张表的数量总和列")
    except Exception as e:
        st.sidebar.error(f"❌ 修复数量总和失败：{e}")

# -------------------- 智能化新增月份/季度 --------------------
st.sidebar.markdown("### 📅 新增数据时间点")
current_year = datetime.now().year
year = st.sidebar.selectbox("选择年份", list(range(current_year - 2, current_year + 2)), index=2)
mode = st.sidebar.radio("时间类型", ["月份", "季度"], horizontal=True)

if mode == "月份":
    month = st.sidebar.selectbox("选择月份", list(range(1, 13)))
    new_sheet_name = f"{year}_{month:02d}"
else:
    quarter = st.sidebar.selectbox("选择季度", ["Q1", "Q2", "Q3", "Q4"])
    new_sheet_name = f"{year}_{quarter}"

if st.sidebar.button("创建新的时间点"):
    if new_sheet_name in sheets:
        st.sidebar.error(f"❌ 时间点 {new_sheet_name} 已存在！")
    else:
        try:
            base_df = pd.DataFrame(columns=["明细", "数量总和", "员工", "值", "分组"])
            prev_sheets = sorted([s for s in sheets if "_" in s and s < new_sheet_name])
            if prev_sheets:
                prev_name = prev_sheets[-1]
                base_df = sheet_frames.get(prev_name, base_df).copy()
                st.sidebar.info(f"🔧 已从最近时间点 {prev_name} 自动继承数据")
            else:
                st.sidebar.info("🔧 未找到上期数据，创建空白模板")

            # 写入时指定引擎
            with pd.ExcelWriter(SAVE_FILE, mode="a", engine="openpyxl") as writer:
                base_df.to_excel(writer, sheet_name=new_sheet_name, index=False)

            st.cache_data.clear()
            sheets, sheet_frames = load_sheets(SAVE_FILE)
            st.sidebar.success(f"✅ 已创建新时间点: {new_sheet_name}")

        except Exception as e:
            st.sidebar.error(f"❌ 创建失败：{e}")

# -------------------- 优化后的删除工作表功能 --------------------
st.sidebar.markdown("### 🗑️ 删除时间点")
if sheets:
    sheet_to_delete = st.sidebar.selectbox("选择要删除的时间点", sheets, key="delete_sheet_select")

    if len(sheets) == 1:
        st.sidebar.warning("⚠️ 至少保留一个工作表，无法删除")
    else:
        if "delete_confirm" not in st.session_state:
            st.session_state.delete_confirm = False

        if not st.session_state.delete_confirm:
            if st.sidebar.button("删除选中时间点", key="delete_btn", help="删除后不可恢复"):
                st.session_state.delete_confirm = True
        else:
            st.sidebar.warning(f"⚠️ 确认删除【{sheet_to_delete}】？此操作不可恢复！")
            col1, col2 = st.sidebar.columns(2)
            with col1:
                if st.button("确认删除", key="confirm_delete"):
                    success, msg = delete_sheet_optimized(SAVE_FILE, sheet_to_delete)
                    st.sidebar.warning(msg)
                    if success:
                        st.cache_data.clear()
                        sheets, sheet_frames = load_sheets(SAVE_FILE)
                        st.session_state.delete_confirm = False
                        st.rerun()
            with col2:
                if st.button("取消", key="cancel_delete"):
                    st.session_state.delete_confirm = False

# -------------------- 🧮 一键更新所有数量总和 --------------------
st.sidebar.markdown("### 🔧 数据修复工具")

if st.sidebar.button("🧮 一键更新所有数量总和"):
    try:
        if not os.path.exists(SAVE_FILE):
            st.sidebar.warning("未找到文件 jixiao.xlsx")
        else:
            xls = pd.ExcelFile(SAVE_FILE, engine="openpyxl")
            updated_frames = {}
            for sheet_name in xls.sheet_names:
                df0 = pd.read_excel(xls, sheet_name=sheet_name, engine="openpyxl")
                if "明细" in df0.columns and "值" in df0.columns:
                    sum_df = (
                        df0.groupby("明细", as_index=False)["值"].sum()
                        .rename(columns={"值": "数量总和"})
                    )
                    df0 = df0.drop(columns=["数量总和"], errors="ignore")
                    df0 = df0.merge(sum_df, on="明细", how="left")
                    updated_frames[sheet_name] = df0

            with pd.ExcelWriter(SAVE_FILE, engine="openpyxl") as writer:
                for sheet_name, df0 in updated_frames.items():
                    df0.to_excel(writer, sheet_name=sheet_name, index=False)

            st.cache_data.clear()
            sheets, sheet_frames = load_sheets(SAVE_FILE)
            st.sidebar.success("✅ 所有工作表的数量总和已重新计算并更新！")

    except Exception as e:
        st.sidebar.error(f"❌ 更新失败：{e}")

# -------------------- 时间点选择优化 --------------------
st.sidebar.markdown("### 📋 数据筛选")
years_available = sorted(list({s.split("_")[0] for s in sheets if "_" in s}))
year_choice = st.sidebar.selectbox("筛选年份", ["全部年份"] + years_available)

if year_choice == "全部年份":
    time_candidates = sorted(sheets)
else:
    time_candidates = sorted([s for s in sheets if s.startswith(year_choice)])

if not time_candidates:
    st.warning(f"⚠️ 暂无符合条件的数据，请先创建月份或季度。")
    time_choice = []
else:
    default_choice = time_candidates[:2] if len(time_candidates) >= 2 else time_candidates[:1]
    time_choice = st.sidebar.multiselect("选择时间点（支持跨年份对比）",
                                         time_candidates,
                                         default=default_choice)

# -------------------- 分组选择 --------------------
all_groups = pd.concat(sheet_frames.values())["分组"].dropna().unique().tolist() if sheet_frames else []
selected_groups = st.sidebar.multiselect("选择分组", all_groups, default=all_groups)

# -------------------- 视图选择 --------------------
sections_names = [
    "人员完成任务数量排名",
    "任务对比（堆叠柱状图）",
    "任务-人员热力图"
]
view = st.sidebar.radio("切换视图", ["编辑数据", "大屏轮播", "单页模式", "显示所有视图", "能力分析"])


# -------------------- 数据合并 --------------------
def get_merged_df(keys: List[str], groups: List[str]) -> pd.DataFrame:
    dfs = []
    for k in keys:
        df0 = sheet_frames.get(k)
        if df0 is not None:
            if groups and "分组" in df0.columns:
                df0 = df0[df0["分组"].isin(groups)]
            dfs.append(df0)
    if not dfs:
        st.warning("⚠️ 当前选择没有数据，请检查时间点或分组选择。")
        return pd.DataFrame()
    return pd.concat(dfs, axis=0, ignore_index=True)


df = get_merged_df(time_choice, selected_groups)


# -------------------- 图表函数 --------------------
def chart_total(df0):
    df0 = df0[df0["明细"] != "分数总和"]
    emp_stats = df0.groupby("员工")["值"].sum().sort_values(ascending=False).reset_index()
    fig = go.Figure(go.Bar(
        x=emp_stats["员工"],
        y=emp_stats["值"],
        text=emp_stats["值"],
        textposition="outside",
        hovertemplate="员工: %{x}<br>完成总值: %{y}<extra></extra>"
    ))
    fig.update_layout(template="plotly_dark", xaxis_title="员工", yaxis_title="完成总值")
    return fig


def chart_stack(df0):
    df0 = df0[df0["明细"] != "分数总和"]
    df_pivot = df0.pivot_table(index="明细", columns="员工", values="值", aggfunc="sum", fill_value=0)
    fig = go.Figure()
    for emp in df_pivot.columns:
        fig.add_trace(go.Bar(x=df_pivot.index, y=df_pivot[emp], name=emp))
    fig.update_layout(barmode="stack", template="plotly_dark", xaxis_title="任务", yaxis_title="完成值")
    return fig


def chart_heat(df0):
    df0 = df0[df0["明细"] != "分数总和"]
    tasks = df0["明细"].unique().tolist()
    emps = df0["员工"].unique().tolist()
    data = []
    for i, t in enumerate(tasks):
        for j, e in enumerate(emps):
            v = int(df0[(df0["明细"] == t) & (df0["员工"] == e)]["值"].sum())
            data.append([j, i, v])
    return {
        "backgroundColor": "transparent",
        "tooltip": {"position": "top"},
        "xAxis": {"type": "category", "data": emps, "axisLabel": {"color": "#fff", "rotate": 45}},
        "yAxis": {"type": "category", "data": tasks, "axisLabel": {"color": "#fff"}},
        "visualMap": {"min": 0, "max": max([d[2] for d in data]) if data else 1, "show": True,
                      "inRange": {"color": ["#ff4d4d", "#4caf50"]}, "textStyle": {"color": "#fff"}},
        "series": [{"type": "heatmap", "data": data, "emphasis": {"itemStyle": {"shadowBlur": 10}}}]
    }


# -------------------- 卡片显示 --------------------
def show_cards(df0):
    df0 = df0[df0["明细"] != "分数总和"]
    if df0.empty:
        return

    total_tasks = df0["明细"].nunique()
    total_people = df0["员工"].nunique()
    ps = df0.groupby("员工")["值"].sum()
    top_person = ps.idxmax() if not ps.empty else ""
    avg_score = round(ps.mean(), 1) if not ps.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(
        f"<div class='metric-card'><div class='metric-value'>{total_tasks}</div><div class='metric-label'>任务数</div></div>",
        unsafe_allow_html=True)
    c2.markdown(
        f"<div class='metric-card'><div class='metric-value'>{total_people}</div><div class='metric-label'>人数</div></div>",
        unsafe_allow_html=True)
    c3.markdown(
        f"<div class='metric-card'><div class='metric-value'>{top_person}</div><div class='metric-label'>覆盖率最高</div></div>",
        unsafe_allow_html=True)
    c4.markdown(
        f"<div class='metric-card'><div class='metric-value'>{avg_score}</div><div class='metric-label'>平均数</div></div>",
        unsafe_allow_html=True)
    st.markdown("<hr/>", unsafe_allow_html=True)


# -------------------- 定义鲜艳的颜色列表 --------------------
BRIGHT_COLORS = [
    "#FF0000", "#00FF00", "#0000FF", "#FFA500", "#800080",
    "#00FFFF", "#FFC0CB", "#FFFF00", "#008080", "#FF00FF"
]

# -------------------- 主页面 --------------------
st.title("📊 技能覆盖分析大屏")

if view == "编辑数据":
    if not time_choice:
        st.warning("⚠️ 请在左侧选择时间点（月或季）后再编辑数据")
    elif len(time_choice) > 1:
        st.warning("⚠️ 编辑数据时仅支持选择单个时间点，请重新选择！")
    else:
        show_cards(df)
        st.info("你可以直接编辑下面的表格，修改完成后点击【保存】按钮。")

        sheet_name = time_choice[0]
        try:
            original_df = pd.read_excel(SAVE_FILE, sheet_name=sheet_name, engine="openpyxl")
            edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True)

            if st.button("💾 保存修改到库里"):
                try:
                    if selected_groups and "分组" in original_df.columns:
                        mask = original_df["分组"].isin(selected_groups)
                        original_df = original_df[~mask].reset_index(drop=True)
                        final_df = pd.concat([original_df, edited_df], ignore_index=True)
                    else:
                        final_df = edited_df.copy()

                    if "明细" in final_df.columns and "值" in final_df.columns:
                        sum_df = (
                            final_df.groupby("明细", as_index=False)["值"].sum()
                            .rename(columns={"值": "数量总和"})
                        )
                        final_df = final_df.drop(columns=["数量总和"], errors="ignore")
                        final_df = final_df.merge(sum_df, on="明细", how="left")

                    with pd.ExcelWriter(SAVE_FILE, mode="a", if_sheet_exists="replace", engine="openpyxl") as writer:
                        final_df.to_excel(writer, sheet_name=sheet_name, index=False)

                    st.cache_data.clear()
                    sheets, sheet_frames = load_sheets(SAVE_FILE)
                    st.success(f"✅ 修改已保存到 {SAVE_FILE} ({sheet_name})，仅更新选中分组数据")
                except Exception as e:
                    st.error(f"保存失败：{e}")
        except Exception as e:
            st.error(f"❌ 加载编辑数据失败：{e}")

elif view == "大屏轮播":
    if not time_choice:
        st.warning("⚠️ 请在左侧选择时间点（月或季）后查看大屏轮播")
    else:
        st_autorefresh(interval=10000, key="aut")
        show_cards(df)
        secs = [("完成排名", chart_total(df)),
                ("任务对比", chart_stack(df)),
                ("热力图", chart_heat(df))]
        t, op = secs[int(time.time() / 10) % len(secs)]
        st.subheader(t)
        if isinstance(op, go.Figure):
            st.plotly_chart(op, use_container_width=True)
        else:
            st.markdown('<div class="heatmap-container">', unsafe_allow_html=True)
            st_echarts(op, height=f"{max(600, len(df['明细'].unique()) * 25)}px", theme="dark")
            st.markdown('</div>', unsafe_allow_html=True)

elif view == "单页模式":
    if not time_choice:
        st.warning("⚠️ 请在左侧选择时间点（月或季）后查看单页模式")
    else:
        show_cards(df)
        choice = st.sidebar.selectbox("单页查看", sections_names, index=0)
        mapping = {
            "人员完成任务数量排名": chart_total(df),
            "任务对比（堆叠柱状图）": chart_stack(df),
            "任务-人员热力图": chart_heat(df)
        }
        chart_func = mapping.get(choice, chart_total(df))
        if isinstance(chart_func, go.Figure):
            st.plotly_chart(chart_func, use_container_width=True)
        else:
            st.markdown('<div class="heatmap-container">', unsafe_allow_html=True)
            st_echarts(chart_func, height=f"{max(600, len(df['明细'].unique()) * 25)}px", theme="dark")
            st.markdown('</div>', unsafe_allow_html=True)

elif view == "显示所有视图":
    if not time_choice:
        st.warning("⚠️ 请在左侧选择时间点（月或季）后查看所有视图")
    else:
        show_cards(df)
        charts = [("完成排名", chart_total(df)),
                  ("任务对比（堆叠柱状图）", chart_stack(df)),
                  ("热图", chart_heat(df))]
        for label, f in charts:
            st.subheader(label)
            if isinstance(f, go.Figure):
                st.plotly_chart(f, use_container_width=True)
            else:
                st.markdown('<div class="heatmap-container">', unsafe_allow_html=True)
                st_echarts(f, height=f"{max(600, len(df['明细'].unique()) * 25)}px", theme="dark")
                st.markdown('</div>', unsafe_allow_html=True)

elif view == "能力分析":
    if not time_choice:
        st.warning("⚠️ 请在左侧选择时间点（月或季）后查看能力分析")
    else:
        st.subheader("📈 能力分析")
        employees = df["员工"].unique().tolist()
        selected_emps = st.sidebar.multiselect("选择员工（图1显示）", employees, default=employees)
        tasks = df["明细"].unique().tolist()

        fig1, fig2, fig3 = go.Figure(), go.Figure(), go.Figure()
        sheet_color_map = {}
        for idx, sheet in enumerate(time_choice):
            sheet_color_map[sheet] = BRIGHT_COLORS[idx % len(BRIGHT_COLORS)]

        emp_color_idx = 0
        for sheet in time_choice:
            df_sheet = get_merged_df([sheet], selected_groups)
            df_sheet = df_sheet[df_sheet["明细"] != "分数总和"]
            df_pivot = df_sheet.pivot(index="明细", columns="员工", values="值").fillna(0)

            for emp in selected_emps:
                fig1.add_trace(go.Scatter(
                    x=tasks,
                    y=df_pivot[emp].reindex(tasks, fill_value=0),
                    mode="lines+markers",
                    name=f"{sheet}-{emp}",
                    line=dict(color=BRIGHT_COLORS[emp_color_idx % len(BRIGHT_COLORS)], width=3),
                    marker=dict(size=8)
                ))
                emp_color_idx += 1

            fig2.add_trace(go.Scatter(
                x=tasks,
                y=df_pivot.sum(axis=1).reindex(tasks, fill_value=0),
                mode="lines+markers",
                name=sheet,
                line=dict(color=sheet_color_map[sheet], width=3),
                marker=dict(size=8)
            ))

            fig3.add_trace(go.Bar(
                x=df_pivot.columns,
                y=df_pivot.sum(axis=0),
                name=sheet,
                marker=dict(color=sheet_color_map[sheet]),
                width=0.3,
            ))

        fig1.update_layout(
            title="员工任务完成情况",
            template="plotly_dark",
            font=dict(size=12),
            legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
            height=500
        )

        fig2.update_layout(
            title="任务整体完成度趋势",
            template="plotly_dark",
            font=dict(size=12),
            legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
            height=500
        )

        fig3.update_layout(
            title="员工整体完成度对比",
            template="plotly_dark",
            font=dict(size=12),
            barmode="group",
            bargap=0.25,
            bargroupgap=0.005,
            legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
            height=600,
            xaxis=dict(
                tickangle=45,
                tickfont=dict(size=10)
            ),
            yaxis=dict(
                tickfont=dict(size=10)
            )
        )

        st.plotly_chart(fig1, use_container_width=True)
        st.plotly_chart(fig2, use_container_width=True)

        st.plotly_chart(fig3, use_container_width=True)
